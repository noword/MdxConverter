#!/usr/bin/env python
# -*- coding: utf-8 -*-
import mdict_query
import openpyxl
import json
import argparse
import os
import pdfkit
from bs4 import BeautifulSoup
import sys
from enum import IntEnum
from collections import OrderedDict
from chardet import detect

ADDITIONAL_STYLES = '''
a.lesson {font-size:120%; color: #1a237e; text-decoration: none; cursor: pointer;}
a.lesson:hover {background-color: #e3f2fd}
a.word {color: #1565c0; text-decoration: none; cursor: pointer;}
a.word:hover {background-color: #e3f2fd;}
div.main {width: 100%; height: 100%;}
div.left {width: 150px; overflow: auto; float: left; height: 100%;}
div.right {overflow-y: auto; overflow-x: hidden; padding-left: 10px; height: 100%;}
'''

H1_STYLE = 'color:#FFFFFF; background-color:#003366; padding-left:20px; line-height:initial;'
H2_STYLE = 'color:#CCFFFF; background-color:#336699; padding-left:20px; line-height:initial;'

INVALID_WORDS_FILENAME = 'invalid_words.txt'


class InvalidAction(IntEnum):
    Exit = 0
    Output = 1
    Collect = 2


def open_encoding_file(name):
    encoding = detect(open(name, 'rb').read())['encoding']
    return open(name, encoding=encoding)


def get_words(name):
    ext = os.path.splitext(name)[1].lower()
    return {'.xls': get_words_from_xls,
            '.xlsx': get_words_from_xls,
            '.json': get_words_from_json,
            '.txt': get_words_from_txt,
            }[ext](name)


def get_words_from_json(name):
    return json.load(open_encoding_file(name))


def get_words_from_txt(name):
    result = []
    for line in open_encoding_file(name).readlines():
        line = line.strip()
        if len(line) == 0:
            continue
        if line.startswith('#'):
            result.append({'name': line.strip('#'), 'words': []})
        else:
            if len(result) == 0:
                result.append({'name': 'Words', 'words': []})
            result[-1]['words'].append(line)
    return result


def get_words_from_xls(name):
    wb = openpyxl.load_workbook(name, read_only=True)
    result = []
    for name in wb.sheetnames:
        ws = wb[name]
        words = [row[0].value for row in ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row, max_col=1)]
        words = list(filter(lambda x: x is not None and len(x) > 0, words))
        result.append({"name": name, "words": words})
    return result


def get_css(soup, mdx_path, dictionary):
    css_name = soup.head.link["href"]
    css_path = os.path.join(mdx_path, css_name)
    if os.path.exists(css_path):
        css = open(css_path, 'rb').read()
    elif hasattr(dictionary, '_mdd_db'):
        css_key = dictionary.get_mdd_keys('*' + css_name)[0]
        css = dictionary.mdd_lookup(css_key)[0]
    else:
        css = b''

    return css.decode('utf-8')


def merge_css(soup, mdx_path, dictionary, append_additinal_styles=True):
    css = get_css(soup, mdx_path, dictionary)
    if append_additinal_styles:
        css += ADDITIONAL_STYLES

    soup.head.link.decompose()
    soup.head.append(soup.new_tag('style', type='text/css'))
    soup.head.style.string = css
    return soup


def grab_images(soup, dictionary):
    if not hasattr(dictionary, '_mdd_db'):
        return

    grabed = set()
    for img in soup.find_all('img'):
        src = img['src'].replace('/', '\\')
        if img['src'].startswith('/'):
            img['src'] = img['src'][1:]
        if src in grabed:
            continue
        grabed.add(src)
        if not src.startswith('\\'):
            src = '\\' + src
        imgs = dictionary.mdd_lookup(src.replace('/', '\\'))
        if len(imgs) > 0:
            src = src[1:]
            print('dump', src)
            try:
                os.makedirs(os.path.split(src)[0])
            except BaseException:
                pass
            open(src, "wb").write(imgs[0])


def lookup(dictionary, word):
    word = word.strip()
    definitions = dictionary.mdx_lookup(word)
    if len(definitions) == 0:
        definitions = dictionary.mdx_lookup(word, ignorecase=True)
    if len(definitions) == 0:
        return ''
    definition = definitions[0]
    if definition.startswith('@@@LINK='):
        return dictionary.mdx_lookup(definition.replace('@@@LINK=', '').strip())[0].strip()
    else:
        return definition.strip()


def verify_words(dictionary, lessons):
    for lesson in lessons:
        print(lesson['name'])
        for word in lesson['words']:
            print('\t', word)
            lookup(dictionary, word)


def mdx2html(mdx_name, input_name, output_name, invalid_action=InvalidAction.Collect, with_toc=True):
    dictionary = mdict_query.IndexBuilder(mdx_name)
    lessons = get_words(input_name)

    right_soup = BeautifulSoup('<body style="font-family:Arial Unicode MS;"><div class="right"></div></body>', 'lxml')
    left_soup = BeautifulSoup('<div class="left"></div>', 'lxml')
    invalid_words = OrderedDict()

    for i, lesson in enumerate(lessons):
        print(lesson['name'])
        h1 = right_soup.new_tag('h1', id='lesson_' + lesson['name'], style=H1_STYLE)
        h1.string = lesson['name']
        right_soup.div.append(h1)

        a = left_soup.new_tag('a', href='#lesson_' + lesson['name'], **{'class': 'lesson'})
        a.string = lesson['name']
        left_soup.div.append(a)
        left_soup.div.append(left_soup.new_tag('br'))

        for j, word in enumerate(lesson['words']):
            print('\t', word)
            result = lookup(dictionary, word)
            if len(result) == 0:  # not found
                print(f'WARNING: "{word}" not found', file=sys.stderr)
                if invalid_action == InvalidAction.Exit:
                    print('*** Exit now. Do nothing. ***')
                    sys.exit()
                elif invalid_action == InvalidAction.Output:
                    result = f'<span><b>WARNING:</b> "{word}" not found</span>'
                else:  # invalid_action == InvalidAction.Collect
                    if lesson['name'] in invalid_words:
                        invalid_words[lesson['name']].append(word)
                    else:
                        invalid_words[lesson['name']] = [word, ]
                    continue
            definition = BeautifulSoup(result, 'lxml')
            if i == j == 0:
                if definition.head is None:
                    right_soup.html.insert_before(right_soup.new_tag('head'))
                else:
                    right_soup.html.insert_before(definition.head)
                right_soup.head.append(right_soup.new_tag('meta', charset='utf-8'))

            h2 = right_soup.new_tag('h2', id='word_' + word, style=H2_STYLE)
            h2.string = word
            right_soup.div.append(h2)
            right_soup.div.append(definition.body)

            a = left_soup.new_tag('a', href='#word_' + word, **{'class': 'word'})
            a.string = word
            left_soup.div.append(a)
            left_soup.div.append(left_soup.new_tag('br'))

        left_soup.div.append(left_soup.new_tag('br'))

    if with_toc:
        main_div = right_soup.new_tag('div', **{'class': 'main'})
        right_soup.div.wrap(main_div)
        right_soup.div.insert_before(left_soup.div)

    right_soup = merge_css(right_soup, os.path.split(mdx_name)[0], dictionary, with_toc)
    grab_images(right_soup, dictionary)

    html = right_soup.prettify().encode('utf-8')
    html = html.replace(b'<body>', b'').replace(b'</body>', b'', html.count(b'</body>') - 1)
    open(output_name, "wb").write(html)

    if len(invalid_words) > 0:
        with open(INVALID_WORDS_FILENAME, 'w') as fp:
            for lesson, words in invalid_words.items():
                fp.write(f'#{lesson}\n')
                for word in words:
                    fp.write(word + '\n')


def mdx2pdf(mdx_name, input_name, output_name, invalid_action=InvalidAction.Collect):
    TEMP_FILE = "temp.html"
    mdx2html(mdx_name, input_name, TEMP_FILE, invalid_action, False)
    pdfkit.from_file(TEMP_FILE, output_name)
    os.remove(TEMP_FILE)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(formatter_class=argparse.RawTextHelpFormatter,)
    parser.add_argument('mdx_name', action='store', nargs=1)
    parser.add_argument('input_name', action='store', nargs=1)
    parser.add_argument('output_name', action='store', nargs="?")
    parser.add_argument('--type', action='store', choices=["pdf", "html"], nargs="?")
    parser.add_argument('--invalid', action='store', type=int, default=2, choices=[0, 1, 2],
                        help='action for meeting invalid words\n'
                        '0: exit immediately\n'
                        '1: output warnning message to pdf/html\n'
                        '2: collect them to invalid_words.txt (default)')
    args = parser.parse_args()

    mdx_name = args.mdx_name[0]
    input_name = args.input_name[0]

    if args.output_name is None:
        if args.type is None:
            raise EnvironmentError('You must choose a file name or a file type')
        else:
            output_name = os.path.split(input_name)[1]
            output_name = os.path.splitext(output_name)[0]
            output_name += "." + args.type
    else:
        output_name = args.output_name

    if args.type is None:
        args.type = os.path.splitext(output_name)[1][1:]

    {
        'pdf': mdx2pdf,
        'html': mdx2html,
    }[args.type.lower()](mdx_name, input_name, output_name, args.invalid)
